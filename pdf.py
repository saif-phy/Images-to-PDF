#Latest
#4 Sept, 2023
#miXAFS log and fig files to PDF


import PySimpleGUI as sg
import os
from openpyxl import load_workbook
from PIL import Image
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from math import floor

def resize_image(image_path, max_width, max_height):
    img = Image.open(image_path)
    img.thumbnail((max_width, max_height), Image.LANCZOS)
    return img

def create_pdf(images, output_file, page_size):
    c = canvas.Canvas(output_file, pagesize=page_size)

    # Calculate the dimensions for each image cell
    page_width, page_height = page_size
    num_images_per_page = 5
    cell_width = floor(page_width)
    cell_height = floor(page_height / num_images_per_page)

    current_image = 0
    current_y = page_height

    for image in images:
        img = resize_image(image, cell_width, cell_height)
        x = 0
        y = current_y - cell_height
        c.drawImage(image, x, y, cell_width, cell_height)

        current_image += 1
        if current_image >= num_images_per_page:
            c.showPage()  # Start a new page
            current_image = 0
            current_y = page_height

        else:
            current_y -= cell_height

    if current_image != 0:
        c.showPage()  # Finish the last page if it's not full

    c.save()

def main():
    sg.theme("DarkGrey5")

    layout = [
        [sg.Text("Select an Excel file (.xlsx):")],
        [sg.InputText(key="xlsx_file"), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
        [sg.Text("Choose file type for images:"), sg.Drop(values=(".jpg", ".png"), key="file_type")],
        [sg.Text("Select the folder containing image files:")],
        [sg.InputText(key="image_folder"), sg.FolderBrowse()],
        [sg.Text("Output PDF File:")],
        [sg.InputText(key="output_file"), sg.FileSaveAs(file_types=(("PDF Files", "*.pdf"),))],
        [sg.Button("Generate PDF"), sg.Button("Exit")],
    ]

    window = sg.Window("pypdf ", layout)

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == "Exit":
            break

        if event == "Generate PDF":
            xlsx_file = values["xlsx_file"]
            file_type = values["file_type"]
            image_folder = values["image_folder"]
            output_file = values["output_file"]

            if not xlsx_file or not file_type or not image_folder or not output_file:
                sg.popup_error("Please fill in all fields.")
                continue

            try:
                wb = load_workbook(xlsx_file)
                sheet = wb.worksheets[0]
                header_name = "LogName"
                # Find the column index that corresponds to the header name
                header_row = sheet[1]  # Assuming headers are in the first row
                column_index = None
                for cell in header_row:
                    if cell.value == header_name:
                        column_index = cell.column
                        break
                images = []
                for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                    for cell in row:
                        if cell.value is not None:
                            image_filename = os.path.join(image_folder, str(cell.value) + file_type)
                            if os.path.exists(image_filename):
                                images.append(image_filename)

                if not images:
                    sg.popup_error("No matching image files found.")
                else:
                    create_pdf(images, output_file, page_size=A4)
                    sg.popup("PDF generated successfully!")

            except Exception as e:
                sg.popup_error(f"An error occurred: {str(e)}")

    window.close()

if __name__ == "__main__":
    main()
